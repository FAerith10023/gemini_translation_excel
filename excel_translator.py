#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 翻译器 - 使用 Gemini AI API
自动检测并翻译 Excel 表格中的中文内容
"""

import re
import json
import time
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from google import genai
from typing import List, Dict, Tuple, Optional
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelTranslator:
    def __init__(self, api_key: str):
        """
        初始化翻译器
        
        Args:
            api_key: Gemini API 密钥
        """
        self.client = genai.Client(api_key=api_key)
        self.chinese_pattern = re.compile(r'[\u4e00-\u9fff]+')
        self.terminology_dict = {}  # 术语库字典
        
    def load_terminology(self, terminology_file: str) -> Dict:
        """
        加载术语库文件
        
        Args:
            terminology_file: 术语库文件路径
            
        Returns:
            Dict: 术语库字典 {中文: 英文}
        """
        try:
            logger.info(f"正在加载术语库: {terminology_file}")
            
            # 读取Excel术语库文件
            df = pd.read_excel(terminology_file)
            
            # 假设术语库文件格式为: 第一列中文，第二列英文
            # 可以根据实际文件格式调整
            if len(df.columns) >= 2:
                chinese_col = df.columns[0]
                english_col = df.columns[1]
                
                terminology_dict = {}
                for _, row in df.iterrows():
                    chinese_term = str(row[chinese_col]).strip() if pd.notna(row[chinese_col]) else ""
                    english_term = str(row[english_col]).strip() if pd.notna(row[english_col]) else ""
                    
                    if chinese_term and english_term:
                        terminology_dict[chinese_term] = english_term
                
                logger.info(f"成功加载 {len(terminology_dict)} 个术语对")
                self.terminology_dict = terminology_dict
                return terminology_dict
            else:
                logger.error("术语库文件格式不正确，需要至少两列（中文和英文）")
                return {}
                
        except Exception as e:
            logger.error(f"加载术语库失败: {str(e)}")
            return {}
    
    def apply_terminology_matching(self, input_file: str, output_file: str, terminology_file: str = None) -> int:
        """
        应用术语库匹配，替换精确匹配的术语
        
        Args:
            input_file: 输入Excel文件路径
            output_file: 输出Excel文件路径
            terminology_file: 术语库文件路径，如果为None则使用默认的terminology_sample.xlsx
            
        Returns:
            int: 替换的术语数量
        """
        try:
            # 加载术语库
            if terminology_file is None:
                terminology_file = "terminology_sample.xlsx"
            
            terminology_dict = self.load_terminology(terminology_file)
            if not terminology_dict:
                logger.warning("术语库为空或加载失败")
                return 0
            
            logger.info(f"开始术语库匹配处理: {input_file}")
            
            # 加载Excel文件
            workbook = load_workbook(input_file)
            replacement_count = 0
            
            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                logger.info(f"处理工作表: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # 获取合并单元格信息
                merged_cells_info = self.extract_merged_cells_info(worksheet)
                
                # 遍历所有单元格
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            cell_value = str(cell.value).strip()
                            
                            # 检查是否在术语库中有精确匹配
                            if cell_value in terminology_dict:
                                old_value = cell_value
                                new_value = terminology_dict[cell_value]
                                
                                # 检查是否为合并单元格
                                cell_coord = cell.coordinate
                                if cell_coord in merged_cells_info:
                                    merged_info = merged_cells_info[cell_coord]
                                    # 只在主单元格更新
                                    if cell_coord == merged_info['master_cell']:
                                        cell.value = new_value
                                        replacement_count += 1
                                        logger.info(f"替换术语 [{sheet_name}]{cell_coord}: '{old_value}' -> '{new_value}'")
                                else:
                                    cell.value = new_value
                                    replacement_count += 1
                                    logger.info(f"替换术语 [{sheet_name}]{cell_coord}: '{old_value}' -> '{new_value}'")
            
            # 保存文件
            workbook.save(output_file)
            workbook.close()
            
            logger.info(f"术语库匹配完成，共替换 {replacement_count} 个术语，结果保存到: {output_file}")
            return replacement_count
            
        except Exception as e:
            logger.error(f"术语库匹配过程中出现错误: {str(e)}")
            raise
    
    def contains_chinese(self, text: str) -> bool:
        """
        检查文本是否包含中文字符
        
        Args:
            text: 要检查的文本
            
        Returns:
            bool: 是否包含中文
        """
        if not isinstance(text, str):
            return False
        return bool(self.chinese_pattern.search(text))
    
    def extract_merged_cells_info(self, worksheet) -> Dict:
        """
        提取合并单元格信息
        
        Args:
            worksheet: openpyxl worksheet 对象
            
        Returns:
            Dict: 合并单元格信息
        """
        merged_cells_info = {}
        for merged_range in worksheet.merged_cells.ranges:
            # 获取合并区域的所有单元格
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_coord = f"{get_column_letter(col)}{row}"
                    merged_cells_info[cell_coord] = {
                        'master_cell': f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}",
                        'range': str(merged_range)
                    }
        return merged_cells_info
    
    def extract_chinese_content(self, file_path: str) -> Dict:
        """
        提取 Excel 文件中所有包含中文的单元格内容
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            Dict: 包含位置和内容的字典
        """
        logger.info(f"正在分析文件: {file_path}")
        
        workbook = load_workbook(file_path)
        chinese_content = {}
        
        for sheet_name in workbook.sheetnames:
            logger.info(f"处理工作表: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # 获取合并单元格信息
            merged_cells_info = self.extract_merged_cells_info(worksheet)
            
            sheet_chinese_content = {}
            
            # 遍历所有单元格
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and self.contains_chinese(str(cell.value)):
                        cell_coord = cell.coordinate
                        
                        # 检查是否为合并单元格
                        is_merged = cell_coord in merged_cells_info
                        
                        sheet_chinese_content[cell_coord] = {
                            'content': str(cell.value),
                            'row': cell.row,
                            'column': cell.column,
                            'is_merged': is_merged,
                            'merged_info': merged_cells_info.get(cell_coord, None)
                        }
            
            if sheet_chinese_content:
                chinese_content[sheet_name] = sheet_chinese_content
        
        workbook.close()
        logger.info(f"找到 {sum(len(content) for content in chinese_content.values())} 个包含中文的单元格")
        return chinese_content
    
    def prepare_translation_batch(self, chinese_content: Dict, keywords: str = "") -> List[Dict]:
        """
        准备翻译批次，将内容按工作表分组
        
        Args:
            chinese_content: 提取的中文内容
            keywords: 专业领域关键词
            
        Returns:
            List[Dict]: 准备好的翻译批次
        """
        translation_batches = []
        
        for sheet_name, content in chinese_content.items():
            if not content:
                continue
                
            # 构建翻译请求
            texts_to_translate = []
            cell_mapping = []
            
            for cell_coord, cell_info in content.items():
                texts_to_translate.append(cell_info['content'])
                cell_mapping.append({
                    'coord': cell_coord,
                    'original': cell_info['content'],
                    'info': cell_info
                })
            
            batch = {
                'sheet_name': sheet_name,
                'texts': texts_to_translate,
                'mapping': cell_mapping,
                'keywords': keywords
            }
            
            translation_batches.append(batch)
        
        return translation_batches
    
    def translate_batch(self, batch: Dict) -> Dict:
        """
        翻译一个批次的内容
        
        Args:
            batch: 批次数据
            
        Returns:
            Dict: 翻译结果
        """
        sheet_name = batch['sheet_name']
        texts = batch['texts']
        keywords = batch.get('keywords', '')
        
        logger.info(f"正在翻译工作表 '{sheet_name}' 中的 {len(texts)} 个文本")
        
        # 构建提示词
        prompt_parts = []
        
        if keywords:
            prompt_parts.append(f"专业领域关键词: {keywords}")
            prompt_parts.append("请根据专业领域进行准确翻译。")
        
        prompt_parts.extend([
            "请将以下中文文本翻译成英文，保持原意和专业性。",
            "请按照输入的顺序返回翻译结果，每行一个翻译结果。",
            "只返回翻译结果，不要添加编号或其他格式。",
            "",
            "待翻译文本:"
        ])
        
        # 添加所有待翻译的文本
        for i, text in enumerate(texts, 1):
            prompt_parts.append(f"{i}. {text}")
        
        prompt = "\n".join(prompt_parts)
        
        try:
            # 调用 Gemini API
            response = self.client.models.generate_content(
                model="gemini-2.0-flash",
                contents=prompt
            )
            
            if response.text:
                # 解析翻译结果
                translated_lines = [line.strip() for line in response.text.strip().split('\n') if line.strip()]
                
                # 确保翻译结果数量与原文本数量匹配
                if len(translated_lines) == len(texts):
                    translations = translated_lines
                else:
                    logger.warning(f"翻译结果数量不匹配: 期望 {len(texts)}, 实际 {len(translated_lines)}")
                    # 如果数量不匹配，尝试逐个翻译
                    translations = self.translate_individually(texts, keywords)
            else:
                logger.error("API 返回空响应")
                translations = self.translate_individually(texts, keywords)
                
        except Exception as e:
            logger.error(f"批量翻译失败: {str(e)}")
            translations = self.translate_individually(texts, keywords)
        
        # 构建结果
        result = {
            'sheet_name': sheet_name,
            'translations': []
        }
        
        for i, (original_info, translation) in enumerate(zip(batch['mapping'], translations)):
            result['translations'].append({
                'coord': original_info['coord'],
                'original': original_info['original'],
                'translation': translation,
                'info': original_info['info']
            })
        
        return result
    
    def translate_individually(self, texts: List[str], keywords: str = "") -> List[str]:
        """
        逐个翻译文本（备用方案）
        
        Args:
            texts: 待翻译文本列表
            keywords: 专业领域关键词
            
        Returns:
            List[str]: 翻译结果列表
        """
        logger.info("使用逐个翻译模式")
        translations = []
        
        for text in texts:
            try:
                prompt_parts = []
                if keywords:
                    prompt_parts.append(f"专业领域关键词: {keywords}")
                
                prompt_parts.extend([
                    "请将以下中文翻译成英文，保持原意和专业性：",
                    text
                ])
                
                prompt = "\n".join(prompt_parts)
                
                response = self.client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=prompt
                )
                
                if response.text:
                    translations.append(response.text.strip())
                else:
                    translations.append(f"[翻译失败: {text}]")
                    
                # 添加延时避免API限制
                time.sleep(0.5)
                
            except Exception as e:
                logger.error(f"翻译文本 '{text}' 时出错: {str(e)}")
                translations.append(f"[翻译失败: {text}]")
        
        return translations
    
    def apply_translations(self, file_path: str, translation_results: List[Dict], output_path: str):
        """
        将翻译结果应用到 Excel 文件
        
        Args:
            file_path: 源文件路径
            translation_results: 翻译结果
            output_path: 输出文件路径
        """
        logger.info("正在应用翻译结果到文件")
        
        # 加载原始文件
        workbook = load_workbook(file_path)
        
        for result in translation_results:
            sheet_name = result['sheet_name']
            
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"工作表 '{sheet_name}' 不存在，跳过")
                continue
                
            worksheet = workbook[sheet_name]
            
            for translation_info in result['translations']:
                coord = translation_info['coord']
                translation = translation_info['translation']
                original_info = translation_info['info']
                
                try:
                    # 获取单元格
                    cell = worksheet[coord]
                    
                    # 处理合并单元格的情况
                    if original_info['is_merged']:
                        merged_info = original_info['merged_info']
                        logger.info(f"处理合并单元格 {coord} (范围: {merged_info['range']})")
                        
                        # 只更新主单元格
                        if coord == merged_info['master_cell']:
                            cell.value = translation
                    else:
                        cell.value = translation
                        
                    logger.debug(f"已更新单元格 {coord}: {translation}")
                    
                except Exception as e:
                    logger.error(f"更新单元格 {coord} 时出错: {str(e)}")
        
        # 保存文件
        workbook.save(output_path)
        workbook.close()
        logger.info(f"翻译完成，结果已保存到: {output_path}")
    
    def translate_excel(self, input_file: str, output_file: str, keywords: str = ""):
        """
        翻译整个 Excel 文件
        
        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径
            keywords: 专业领域关键词
        """
        try:
            # 1. 提取中文内容
            chinese_content = self.extract_chinese_content(input_file)
            
            if not chinese_content:
                logger.info("未找到包含中文的单元格")
                return
            
            # 2. 一次性翻译所有中文内容
            translation_result = self.translate_all_content(chinese_content, keywords)
            
            # 3. 应用翻译结果
            self.apply_all_translations(input_file, translation_result, output_file)
            
            logger.info("Excel 翻译完成!")
            
        except Exception as e:
            logger.error(f"翻译过程中出现错误: {str(e)}")
            raise

    def translate_all_content(self, chinese_content: Dict, keywords: str = "") -> Dict:
        """
        一次性翻译所有中文内容
        
        Args:
            chinese_content: 提取的中文内容
            keywords: 专业领域关键词
            
        Returns:
            Dict: 翻译结果
        """
        logger.info("开始一次性翻译所有中文内容")
        
        # 收集所有需要翻译的文本和对应位置信息
        all_texts = []
        text_mapping = []
        
        for sheet_name, content in chinese_content.items():
            for cell_coord, cell_info in content.items():
                all_texts.append(cell_info['content'])
                text_mapping.append({
                    'sheet_name': sheet_name,
                    'coord': cell_coord,
                    'original': cell_info['content'],
                    'info': cell_info
                })
        
        total_texts = len(all_texts)
        logger.info(f"共需要翻译 {total_texts} 个文本")
        
        if total_texts == 0:
            return {'translations': []}
        
        # 构建提示词
        prompt_parts = []
        
        if keywords:
            prompt_parts.append(f"专业领域关键词: {keywords}")
            prompt_parts.append("请根据专业领域进行准确翻译。")
        
        prompt_parts.extend([
            "请将以下中文文本翻译成英文，保持原意和专业性。",
            "请按照输入的顺序返回翻译结果，每行一个翻译结果。",
            "只返回翻译结果，不要添加编号或其他格式。",
            "注意：输入可能包含来自不同工作表的内容，请逐一翻译。",
            "",
            "待翻译文本:"
        ])
        
        # 添加所有待翻译的文本
        for i, text in enumerate(all_texts, 1):
            prompt_parts.append(f"{i}. {text}")
        
        prompt = "\n".join(prompt_parts)
        
        try:
            # 调用 Gemini API 一次性翻译所有内容
            logger.info("正在调用 Gemini API 进行翻译...")
            response = self.client.models.generate_content(
                model="gemini-2.0-flash",
                contents=prompt
            )
            
            if response.text:
                # 解析翻译结果
                translated_lines = [line.strip() for line in response.text.strip().split('\n') if line.strip()]
                
                # 确保翻译结果数量与原文本数量匹配
                if len(translated_lines) == len(all_texts):
                    translations = translated_lines
                    logger.info("API 翻译成功，结果数量匹配")
                else:
                    logger.warning(f"翻译结果数量不匹配: 期望 {len(all_texts)}, 实际 {len(translated_lines)}")
                    # 如果数量不匹配，尝试逐个翻译
                    logger.info("切换到逐个翻译模式")
                    translations = self.translate_individually(all_texts, keywords)
            else:
                logger.error("API 返回空响应，切换到逐个翻译模式")
                translations = self.translate_individually(all_texts, keywords)
                
        except Exception as e:
            logger.error(f"一次性翻译失败: {str(e)}，切换到逐个翻译模式")
            translations = self.translate_individually(all_texts, keywords)
        
        # 构建结果字典
        result = {
            'translations': []
        }
        
        for i, (mapping_info, translation) in enumerate(zip(text_mapping, translations)):
            result['translations'].append({
                'sheet_name': mapping_info['sheet_name'],
                'coord': mapping_info['coord'],
                'original': mapping_info['original'],
                'translation': translation,
                'info': mapping_info['info']
            })
        
        logger.info(f"翻译完成，共处理 {len(result['translations'])} 个文本")
        return result

    def apply_all_translations(self, file_path: str, translation_result: Dict, output_path: str):
        """
        将所有翻译结果应用到 Excel 文件
        
        Args:
            file_path: 源文件路径
            translation_result: 翻译结果
            output_path: 输出文件路径
        """
        logger.info("正在应用翻译结果到文件")
        
        # 加载原始文件
        workbook = load_workbook(file_path)
        
        # 按工作表分组翻译结果
        sheet_translations = {}
        for trans in translation_result['translations']:
            sheet_name = trans['sheet_name']
            if sheet_name not in sheet_translations:
                sheet_translations[sheet_name] = []
            sheet_translations[sheet_name].append(trans)
        
        # 应用翻译结果
        for sheet_name, translations in sheet_translations.items():
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"工作表 '{sheet_name}' 不存在，跳过")
                continue
                
            worksheet = workbook[sheet_name]
            logger.info(f"正在处理工作表 '{sheet_name}'，共 {len(translations)} 个翻译")
            
            for translation_info in translations:
                coord = translation_info['coord']
                translation = translation_info['translation']
                original_info = translation_info['info']
                
                try:
                    # 获取单元格
                    cell = worksheet[coord]
                    
                    # 处理合并单元格的情况
                    if original_info['is_merged']:
                        merged_info = original_info['merged_info']
                        logger.debug(f"处理合并单元格 {coord} (范围: {merged_info['range']})")
                        
                        # 只更新主单元格
                        if coord == merged_info['master_cell']:
                            cell.value = translation
                    else:
                        cell.value = translation
                        
                    logger.debug(f"已更新单元格 [{sheet_name}]{coord}: {translation}")
                    
                except Exception as e:
                    logger.error(f"更新单元格 [{sheet_name}]{coord} 时出错: {str(e)}")
        
        # 保存文件
        workbook.save(output_path)
        workbook.close()
        logger.info(f"翻译完成，结果已保存到: {output_path}")


def main():
    """主函数"""
    print("=== Excel 中文翻译器 (使用 Gemini AI) ===\n")
    
    # 获取用户输入
    api_key = input("请输入您的 Gemini API 密钥: ").strip()
    if not api_key:
        print("错误: API 密钥不能为空")
        return
    
    input_file = input("请输入要翻译的 Excel 文件路径: ").strip()
    if not input_file:
        print("错误: 文件路径不能为空")
        return
    
    # 生成输出文件名
    if '.' in input_file:
        base_name = input_file.rsplit('.', 1)[0]
        extension = input_file.rsplit('.', 1)[1]
        output_file = f"{base_name}_translated.{extension}"
    else:
        output_file = f"{input_file}_translated.xlsx"
    
    print(f"翻译结果将保存到: {output_file}")
    
    # 获取专业领域关键词（可选）
    keywords = input("请输入专业领域关键词（可选，如：医学、法律、技术等）: ").strip()
    
    try:
        # 创建翻译器实例
        translator = ExcelTranslator(api_key)
        
        # 开始翻译
        print("\n开始翻译...")
        translator.translate_excel(input_file, output_file, keywords)
        
        print(f"\n翻译完成! 结果已保存到: {output_file}")
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 '{input_file}'")
    except PermissionError:
        print(f"错误: 没有权限访问文件 '{input_file}' 或 '{output_file}'")
    except Exception as e:
        print(f"错误: {str(e)}")


if __name__ == "__main__":
    main() 