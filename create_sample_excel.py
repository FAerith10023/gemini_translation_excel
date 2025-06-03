#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建示例 Excel 文件用于测试翻译功能
包含中文内容、合并单元格等情况
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_sample_excel():
    """创建包含中文内容的示例 Excel 文件"""
    
    # 创建工作簿
    wb = Workbook()
    
    # 第一个工作表 - 产品信息
    ws1 = wb.active
    ws1.title = "产品信息"
    
    # 设置标题行（合并单元格）
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "产品信息表"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 设置表头
    headers = ["产品名称", "产品描述", "价格", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # 添加产品数据
    products = [
        ["智能手机", "高性能智能手机，支持5G网络", "￥3999", "热销产品"],
        ["笔记本电脑", "轻薄便携笔记本电脑，适合办公", "￥6999", "商务首选"],
        ["无线耳机", "蓝牙无线耳机，降噪功能", "￥599", "音质优秀"],
        ["平板电脑", "10英寸高清屏幕平板电脑", "￥2999", "娱乐学习两用"],
        ["智能手表", "健康监测智能手表", "￥1599", "运动必备"]
    ]
    
    for row, product in enumerate(products, 3):
        for col, value in enumerate(product, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # 合并备注列的某些单元格
    ws1.merge_cells('D6:D7')
    ws1['D6'] = "限时优惠活动中"
    
    # 第二个工作表 - 公司信息
    ws2 = wb.create_sheet("公司信息")
    
    # 公司信息数据
    company_info = [
        ["公司名称", "科技创新有限公司"],
        ["成立时间", "2020年3月"],
        ["主营业务", "电子产品研发与销售"],
        ["员工人数", "150人"],
        ["年营业额", "5000万元"],
        ["发展目标", "打造行业领先的科技企业"],
        ["核心价值", "创新、品质、服务"],
        ["联系地址", "北京市海淀区中关村大街123号"]
    ]
    
    # 合并标题
    ws2.merge_cells('A1:B1')
    ws2['A1'] = "公司基本信息"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加公司信息
    for row, (key, value) in enumerate(company_info, 2):
        ws2.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)
    
    # 第三个工作表 - 技术规格
    ws3 = wb.create_sheet("技术规格")
    
    # 技术规格数据
    tech_specs = [
        ["处理器", "八核处理器", "高性能芯片"],
        ["内存", "8GB RAM", "流畅运行"],
        ["存储", "256GB SSD", "快速存储"],
        ["显示", "15.6英寸高清屏", "色彩鲜艳"],
        ["电池", "长续航电池", "全天使用"],
        ["系统", "最新操作系统", "稳定可靠"],
        ["网络", "Wi-Fi 6 + 蓝牙5.0", "连接稳定"],
        ["重量", "1.8kg", "轻便携带"]
    ]
    
    # 设置标题
    ws3.merge_cells('A1:C1')
    ws3['A1'] = "技术规格参数表"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 设置子标题
    headers = ["参数名称", "参数值", "说明"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # 添加技术规格数据
    for row, spec in enumerate(tech_specs, 3):
        for col, value in enumerate(spec, 1):
            ws3.cell(row=row, column=col, value=value)
    
    # 调整列宽
    for ws in [ws1, ws2, ws3]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # 保存文件
    filename = "sample_chinese_excel.xlsx"
    wb.save(filename)
    print(f"示例 Excel 文件已创建: {filename}")
    
    return filename

if __name__ == "__main__":
    create_sample_excel() 