import pandas as pd
from openpyxl import load_workbook
import re

def has_chinese(text):
    """检查文本是否包含中文字符"""
    if not isinstance(text, str):
        return False
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
    return bool(chinese_pattern.search(text))

def show_translation_comparison():
    """显示翻译前后的对比"""
    try:
        # 加载原文件和翻译后的文件
        input_wb = load_workbook('input.xlsx')
        output_wb = load_workbook('output.xlsx')
        
        input_ws = input_wb.active
        output_ws = output_wb.active
        
        print("=== Excel 中文→英文翻译结果对比 ===\n")
        
        # 扫描A1:O20区域
        comparisons = []
        for row in range(1, 21):
            for col in range(1, 16):
                input_cell = input_ws.cell(row=row, column=col)
                output_cell = output_ws.cell(row=row, column=col)
                
                if input_cell.value and has_chinese(str(input_cell.value)):
                    original = str(input_cell.value).strip()
                    translated = str(output_cell.value).strip() if output_cell.value else "未翻译"
                    
                    # 转换行列为Excel格式（如A1, B2等）
                    col_letter = chr(ord('A') + col - 1) if col <= 26 else f"{chr(ord('A') + (col-1)//26 - 1)}{chr(ord('A') + (col-1)%26)}"
                    cell_ref = f"{col_letter}{row}"
                    
                    comparisons.append({
                        'cell': cell_ref,
                        'original': original,
                        'translated': translated
                    })
        
        # 显示对比结果
        print(f"共翻译了 {len(comparisons)} 个单元格:\n")
        
        for i, comp in enumerate(comparisons, 1):
            print(f"{i:2d}. 单元格 {comp['cell']}:")
            print(f"    原文: {comp['original'][:80]}{'...' if len(comp['original']) > 80 else ''}")
            print(f"    译文: {comp['translated'][:80]}{'...' if len(comp['translated']) > 80 else ''}")
            print()
            
            # 每显示5个就暂停一下
            if i % 5 == 0 and i < len(comparisons):
                input("按回车键继续查看...")
                print()
        
        print("=== 翻译完成 ===")
        print(f"输入文件: input.xlsx")
        print(f"输出文件: output.xlsx")
        print(f"总计翻译: {len(comparisons)} 个单元格")
        
    except FileNotFoundError as e:
        print(f"文件未找到: {e}")
    except Exception as e:
        print(f"出现错误: {e}")

if __name__ == "__main__":
    show_translation_comparison() 