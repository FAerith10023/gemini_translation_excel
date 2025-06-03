#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 翻译器 Web 应用
使用 Flask 构建的 Web 前端界面
"""

import os
import time
import uuid
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
from google import genai
from excel_translator import ExcelTranslator
import logging
import shutil

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 创建 Flask 应用
app = Flask(__name__)
app.secret_key = 'excel-translator-secret-key-2024'  # 用于session和flash消息

# 配置文件上传
UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# 确保文件夹存在
Path(UPLOAD_FOLDER).mkdir(exist_ok=True)
Path(DOWNLOAD_FOLDER).mkdir(exist_ok=True)


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def test_gemini_api(api_key):
    """测试 Gemini API 连接"""
    try:
        client = genai.Client(api_key=api_key)
        
        # 发送简单的测试请求
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents="请回复'API连接成功'"
        )
        
        if response.text and "成功" in response.text:
            return True, "API 连接成功"
        else:
            return True, f"API 响应: {response.text[:50]}..."
            
    except Exception as e:
        return False, f"API 连接失败: {str(e)}"


@app.route('/')
def index():
    """主页"""
    return render_template('index.html')


@app.route('/api/test-connection', methods=['POST'])
def test_connection():
    """测试 API 连接"""
    try:
        data = request.get_json()
        api_key = data.get('api_key', '').strip()
        
        if not api_key:
            return jsonify({
                'success': False,
                'message': 'API 密钥不能为空'
            })
        
        success, message = test_gemini_api(api_key)
        
        return jsonify({
            'success': success,
            'message': message
        })
        
    except Exception as e:
        logger.error(f"测试连接时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'测试失败: {str(e)}'
        })


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """文件上传接口"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            })
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            })
        
        if file and allowed_file(file.filename):
            # 生成唯一文件名
            timestamp = int(time.time())
            unique_id = str(uuid.uuid4())[:8]
            filename = secure_filename(file.filename)
            name, ext = os.path.splitext(filename)
            unique_filename = f"{name}_{timestamp}_{unique_id}{ext}"
            
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(filepath)
            
            # 获取文件信息
            file_size = os.path.getsize(filepath)
            
            return jsonify({
                'success': True,
                'message': '文件上传成功',
                'filename': unique_filename,
                'original_name': filename,
                'size': file_size
            })
        else:
            return jsonify({
                'success': False,
                'message': '文件类型不支持，请上传 .xlsx 或 .xls 文件'
            })
            
    except Exception as e:
        logger.error(f"文件上传时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'上传失败: {str(e)}'
        })


@app.route('/api/terminology-match', methods=['POST'])
def terminology_match():
    """术语库匹配接口"""
    try:
        data = request.get_json()
        
        api_key = data.get('api_key', '').strip()
        filename = data.get('filename', '').strip()
        
        # 验证参数
        if not api_key:
            return jsonify({
                'success': False,
                'message': 'API 密钥不能为空'
            })
        
        if not filename:
            return jsonify({
                'success': False,
                'message': '请先上传文件'
            })
        
        # 检查文件是否存在
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(input_path):
            return jsonify({
                'success': False,
                'message': '文件不存在，请重新上传'
            })
        
        # 生成输出文件名
        name, ext = os.path.splitext(filename)
        output_filename = f"{name}_terminology_matched{ext}"
        output_path = os.path.join(app.config['DOWNLOAD_FOLDER'], output_filename)
        
        # 执行术语库匹配
        logger.info(f"开始术语库匹配: {filename}")
        translator = ExcelTranslator(api_key=api_key)
        
        # 应用术语库匹配
        replacement_count = translator.apply_terminology_matching(
            input_file=input_path,
            output_file=output_path
        )
        
        # 同时将匹配后的文件复制到uploads文件夹以供后续翻译使用
        matched_upload_filename = f"{name}_terminology_matched{ext}"
        matched_upload_path = os.path.join(app.config['UPLOAD_FOLDER'], matched_upload_filename)
        shutil.copy2(output_path, matched_upload_path)
        
        return jsonify({
            'success': True,
            'message': f'术语库匹配完成，共替换 {replacement_count} 个术语',
            'download_filename': output_filename,
            'matched_filename': matched_upload_filename,  # 用于后续翻译的文件名
            'replacement_count': replacement_count,
            'file_size': os.path.getsize(output_path)
        })
        
    except Exception as e:
        logger.error(f"术语库匹配时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'术语库匹配失败: {str(e)}'
        })


@app.route('/api/translate', methods=['POST'])
def translate_excel():
    """Excel 翻译接口"""
    try:
        data = request.get_json()
        
        api_key = data.get('api_key', '').strip()
        filename = data.get('filename', '').strip()
        keywords = data.get('keywords', '').strip()
        
        # 验证参数
        if not api_key:
            return jsonify({
                'success': False,
                'message': 'API 密钥不能为空'
            })
        
        if not filename:
            return jsonify({
                'success': False,
                'message': '请先上传文件'
            })
        
        # 检查文件是否存在
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(input_path):
            return jsonify({
                'success': False,
                'message': '文件不存在，请重新上传'
            })
        
        # 生成输出文件名
        name, ext = os.path.splitext(filename)
        output_filename = f"{name}_translated{ext}"
        output_path = os.path.join(app.config['DOWNLOAD_FOLDER'], output_filename)
        
        # 执行翻译
        logger.info(f"开始翻译文件: {filename}")
        translator = ExcelTranslator(api_key=api_key)
        
        # 这里我们需要在后台执行翻译，返回任务ID
        # 为了简化，这里直接执行翻译
        translator.translate_excel(
            input_file=input_path,
            output_file=output_path,
            keywords=keywords
        )
        
        # 检查输出文件是否生成
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            
            return jsonify({
                'success': True,
                'message': '翻译完成',
                'download_filename': output_filename,
                'output_size': file_size
            })
        else:
            return jsonify({
                'success': False,
                'message': '翻译完成但未生成输出文件'
            })
            
    except Exception as e:
        logger.error(f"翻译时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'翻译失败: {str(e)}'
        })


@app.route('/api/download/<filename>')
def download_file(filename):
    """文件下载接口"""
    try:
        file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            flash('文件不存在', 'error')
            return redirect(url_for('index'))
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"文件下载时出错: {str(e)}")
        flash(f'下载失败: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/api/file-info/<filename>')
def get_file_info(filename):
    """获取文件信息"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': '文件不存在'
            })
        
        # 简单分析文件（提取中文单元格数量等）
        try:
            from openpyxl import load_workbook
            import re
            
            workbook = load_workbook(file_path)
            chinese_pattern = re.compile(r'[\u4e00-\u9fff]+')
            
            total_sheets = len(workbook.sheetnames)
            chinese_cells = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if chinese_pattern.search(cell.value):
                                chinese_cells += 1
            
            workbook.close()
            
            return jsonify({
                'success': True,
                'info': {
                    'total_sheets': total_sheets,
                    'chinese_cells': chinese_cells,
                    'sheets': workbook.sheetnames if 'workbook' in locals() else []
                }
            })
            
        except Exception as e:
            return jsonify({
                'success': True,
                'info': {
                    'error': f'无法分析文件: {str(e)}'
                }
            })
            
    except Exception as e:
        logger.error(f"获取文件信息时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取文件信息失败: {str(e)}'
        })


@app.route('/health')
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'healthy',
        'timestamp': time.time()
    })


@app.errorhandler(413)
def too_large(e):
    """文件过大错误处理"""
    return jsonify({
        'success': False,
        'message': f'文件过大，最大允许 {MAX_FILE_SIZE // (1024*1024)} MB'
    }), 413


@app.errorhandler(500)
def internal_error(e):
    """内部服务器错误处理"""
    logger.error(f"内部服务器错误: {str(e)}")
    return jsonify({
        'success': False,
        'message': '服务器内部错误'
    }), 500


if __name__ == '__main__':
    # 开发模式
    app.run(debug=True, host='0.0.0.0', port=5000) 